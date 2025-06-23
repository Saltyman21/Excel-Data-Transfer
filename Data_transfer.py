from openpyxl import load_workbook

wb_source = load_workbook("source.xlsx")
ws_source = wb_source.active

wb_target = load_workbook("template.xlsx")
ws_target = wb_target.active

# Copy values (assuming headers match and data starts at row 2)
for i in range(2, 52):  # 50 rows
    for j in range(1, ws_target.max_column + 1):
        ws_target.cell(row=i, column=j).value = ws_source.cell(row=i, column=j).value

wb_target.save("output.xlsx")